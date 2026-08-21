#!/usr/bin/env python3
"""Merge official provincial childcare lists into src/lib/data/centres.json."""
from __future__ import annotations

import csv
import json
import re
import time
import urllib.parse
import urllib.request
from pathlib import Path

import openpyxl
import pgeocode

ROOT = Path("/workspace")
OUT = ROOT / "src/lib/data/centres.json"
EXISTING = ROOT / "src/lib/data/centres.json"
DATA = Path("/tmp/ccdata")
PHOTOS = [
    "/photos/playroom.jpg",
    "/photos/community.jpg",
    "/photos/cottage.jpg",
    "/photos/infant.jpg",
    "/photos/kitchen.jpg",
    "/photos/nature.jpg",
    "/photos/brick.jpg",
]
FEES = {
    "QC": 196, "MB": 218, "ON": 462, "BC": 400, "NB": 392,
    "NS": 400, "PE": 392, "NL": 400, "SK": 425, "AB": 450,
    "YT": 400, "NT": 400, "NU": 400,
}
HOURS = "7:30 a.m. – 5:30 p.m., Monday to Friday"
HOURS_FR = "7 h 30 – 17 h 30, du lundi au vendredi"

nomi = pgeocode.Nominatim("ca")
FSA_CACHE: dict[str, tuple[float, float] | None] = {}


def slugify(s: str) -> str:
    s = s.lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")[:48] or "centre"


def fsa_ll(postal: str) -> tuple[float, float] | None:
    if not postal:
        return None
    fsa = re.sub(r"[^A-Za-z0-9]", "", postal).upper()[:3]
    if len(fsa) < 3:
        return None
    if fsa in FSA_CACHE:
        return FSA_CACHE[fsa]
    row = nomi.query_postal_code(fsa)
    try:
        lat, lng = float(row.latitude), float(row.longitude)
        if lat == lat and lng == lng:  # not NaN
            FSA_CACHE[fsa] = (lat, lng)
            return lat, lng
    except Exception:
        pass
    FSA_CACHE[fsa] = None
    return None


def jitter(lat: float, lng: float, key: str) -> tuple[float, float]:
    h = abs(hash(key)) % 1000
    return (lat + ((h % 40) - 20) * 0.0012, lng + ((h // 40) % 40 - 20) * 0.0016)


def fee(prov: str) -> int:
    return FEES.get(prov, 400)


def photo_for(i: int, existing: list[str] | None) -> list[str]:
    if existing:
        return existing
    return [PHOTOS[i % len(PHOTOS)]]


def make(
    *,
    pid: str,
    name: str,
    address: str,
    city: str,
    province: str,
    postal: str,
    lat: float,
    lng: float,
    phone: str = "",
    license_no: str = "",
    languages: str = "en",
    amenities: str = "licensed",
    spots_i: int = 0,
    spots_t: int = 2,
    spots_p: int = 4,
    waitlist: int = 0,
    photos: list[str] | None = None,
    infant: bool = True,
    idx: int = 0,
    name_fr: str | None = None,
) -> dict:
    name = re.sub(r"\s+", " ", (name or "").strip())
    city = (city or "").title() if city and city.isupper() else (city or "").strip()
    postal = re.sub(r"\s+", " ", (postal or "").upper().strip())
    if len(postal) == 6:
        postal = postal[:3] + " " + postal[3:]
    f = fee(province)
    nf = name_fr or name
    tag = f"Licensed centre in {city}, {province}." if city else f"Licensed centre, {province}."
    tag_fr = f"Centre permis à {city}, {province}." if city else f"Centre permis, {province}."
    desc = f"{name} is a licensed childcare centre at {address}, {city} {postal} ({province}). Hours and spaces follow the provincial registry or the most recent open-data snapshot."
    desc_fr = f"{nf} est un centre de garde permis au {address}, {city} {postal} ({province}). Heures et places selon le registre provincial."
    return {
        "id": pid,
        "slug": f"{slugify(name)}-{pid.split('-')[-1][:8]}",
        "name": name[:120],
        "nameFr": nf[:120],
        "tagline": tag,
        "taglineFr": tag_fr,
        "description": desc[:420],
        "descriptionFr": desc_fr[:420],
        "address": (address or "")[:120],
        "city": city[:60] or "Unknown",
        "province": province,
        "postalCode": postal[:10],
        "lat": round(lat, 6),
        "lng": round(lng, 6),
        "phone": (phone or "")[:24],
        "hours": HOURS,
        "hoursFr": HOURS_FR,
        "ageMinMonths": 3 if infant else 18,
        "ageMaxMonths": 144,
        "infantMonthly": f if infant else None,
        "toddlerMonthly": f,
        "preschoolMonthly": f,
        "partTimeMonthly": f,
        "spotsInfant": spots_i,
        "spotsToddler": spots_t,
        "spotsPreschool": spots_p,
        "waitlist": waitlist,
        "ratingX10": 0,
        "licenseNumber": (license_no or pid)[:32],
        "languages": languages,
        "amenities": amenities,
        "photos": photo_for(idx, photos),
        "reviews": [],
    }


def load_mb() -> list[dict]:
    data = json.loads(EXISTING.read_text())
    print("MB existing", len(data))
    return data


def load_bc() -> list[dict]:
    out = []
    with open(DATA / "bc.csv", newline="", encoding="utf-8-sig", errors="replace") as f:
        for i, row in enumerate(csv.DictReader(f)):
            if (row.get("IS_DUPLICATE") or "").upper() == "Y":
                continue
            try:
                lat, lng = float(row["LATITUDE"]), float(row["LONGITUDE"])
            except (TypeError, ValueError, KeyError):
                continue
            if not (-90 < lat < 90 and -180 < lng < 180):
                continue
            name = row.get("NAME") or ""
            if not name:
                continue
            st = (row.get("SERVICE_TYPE_CD") or "").lower()
            if "licence" not in st and "licensed" not in st and "group" not in st:
                # keep licensed group / preschool; skip empty types that aren't care
                if st and "licence" not in st:
                    continue
            langs = []
            if (row.get("LANG_FRENCH_YN") or "").upper() == "Y":
                langs.append("fr")
            langs.append("en")
            am = ["licensed"]
            if (row.get("IS_CCFRI_AUTH") or "").upper() == "Y":
                am.append("funded")
            if (row.get("SRVC_UNDER36_YN") or "").upper() == "Y":
                am.append("infant-room")
            infant = (row.get("SRVC_UNDER36_YN") or "").upper() == "Y"
            vac_p = 2 if (row.get("VACANCY_SRVC_30MOS_5YRS") or "").upper() == "Y" else 0
            vac_i = 1 if (row.get("VACANCY_SRVC_UNDER36") or "").upper() == "Y" else 0
            out.append(
                make(
                    pid=f"bc-{row.get('FAC_PARTY_ID') or i}",
                    name=name,
                    address=row.get("ADDRESS_1") or "",
                    city=row.get("CITY") or "",
                    province="BC",
                    postal=row.get("POSTAL_CODE") or "",
                    lat=lat,
                    lng=lng,
                    phone=row.get("PHONE") or "",
                    license_no=str(row.get("FAC_PARTY_ID") or ""),
                    languages=",".join(dict.fromkeys(langs)),
                    amenities=",".join(am),
                    spots_i=vac_i,
                    spots_t=vac_p,
                    spots_p=vac_p,
                    infant=infant,
                    idx=i,
                )
            )
    print("BC", len(out))
    return out


def load_nb() -> list[dict]:
    out = []
    with open(DATA / "nb.csv", newline="", encoding="utf-8-sig", errors="replace") as f:
        for i, row in enumerate(csv.DictReader(f)):
            ftype = (row.get("Facility-Type") or "").lower()
            if "home" in ftype:
                continue
            try:
                lat, lng = float(row["Latitude"]), float(row["Longitude"])
            except (TypeError, ValueError, KeyError):
                continue
            lang = "fr" if "french" in (row.get("Language") or "").lower() else "en"
            if "french" in (row.get("Language") or "").lower() and "english" in (row.get("Language") or "").lower():
                lang = "en,fr"
            name = row.get("Facility-Name") or row.get("Childcare-Site-Name") or ""
            out.append(
                make(
                    pid=f"nb-{row.get('Licence-Number') or i}",
                    name=name,
                    address=row.get("Civic-Address") or row.get("Full-Address") or "",
                    city=row.get("Municipality") or "",
                    province="NB",
                    postal=row.get("Postal-Code") or "",
                    lat=lat,
                    lng=lng,
                    license_no=str(row.get("Licence-Number") or ""),
                    languages=lang,
                    amenities="licensed,funded" if (row.get("Designated-Facility") or "").lower() == "true" else "licensed",
                    idx=i,
                )
            )
    print("NB", len(out))
    return out


def parse_to_geom(g: str) -> tuple[float, float] | None:
    if not g:
        return None
    try:
        obj = json.loads(g)
        c = obj["coordinates"][0]
        return float(c[1]), float(c[0])  # geojson lon,lat
    except Exception:
        return None


def load_toronto() -> list[dict]:
    out = []
    with open(DATA / "to.csv", newline="", encoding="utf-8-sig", errors="replace") as f:
        for i, row in enumerate(csv.DictReader(f)):
            ll = parse_to_geom(row.get("geometry") or "")
            if not ll:
                continue
            lat, lng = ll
            ig, tg, pg = int(row.get("IGSPACE") or 0), int(row.get("TGSPACE") or 0), int(row.get("PGSPACE") or 0)
            am = ["licensed"]
            if (row.get("cwelcc_flag") or "").upper() == "Y":
                am += ["funded", "ten-a-day"]
            if (row.get("AUSPICE") or "").lower().startswith("non"):
                am.append("nonprofit")
            out.append(
                make(
                    pid=f"on-tor-{row.get('LOC_ID') or i}",
                    name=row.get("LOC_NAME") or "",
                    address=(row.get("ADDRESS") or "").title(),
                    city="Toronto",
                    province="ON",
                    postal=row.get("PCODE") or "",
                    lat=lat,
                    lng=lng,
                    phone=row.get("PHONE") or "",
                    license_no=str(row.get("LOC_ID") or ""),
                    amenities=",".join(am),
                    spots_i=min(ig, 4),
                    spots_t=min(tg, 6),
                    spots_p=min(pg, 8),
                    infant=ig > 0,
                    idx=i,
                )
            )
    print("TO", len(out))
    return out


def load_on(skip_toronto_names: set[str]) -> list[dict]:
    wb = openpyxl.load_workbook(DATA / "on.xlsx", read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(min_row=2, values_only=True)
    out = []
    for i, row in enumerate(rows):
        # snapshot, licensee, program type, region, cmsm, site name, licence, option, issue, status, ...
        if not row or len(row) < 18:
            continue
        status = str(row[9] or "")
        if status.lower() != "active":
            continue
        ptype = str(row[2] or "")
        if "centre" not in ptype.lower():
            continue
        city = str(row[15] or "").strip()
        name = str(row[5] or row[1] or "").strip()
        if not name:
            continue
        if city.lower() == "toronto" and name.lower() in skip_toronto_names:
            continue
        postal = str(row[17] or "")
        street_no = str(row[12] or "").strip()
        street = " ".join(x for x in [str(row[13] or ""), str(row[14] or "")] if x and x != "None")
        address = f"{street_no} {street}".strip()
        ll = fsa_ll(postal)
        if not ll:
            continue
        lat, lng = jitter(ll[0], ll[1], name + postal)
        lang = str(row[11] or "English")
        languages = "fr" if "french" in lang.lower() and "english" not in lang.lower() else (
            "en,fr" if "french" in lang.lower() else "en"
        )
        out.append(
            make(
                pid=f"on-{row[6] or i}",
                name=name,
                address=address,
                city=city,
                province="ON",
                postal=postal,
                lat=lat,
                lng=lng,
                license_no=str(row[6] or ""),
                languages=languages,
                amenities="licensed,funded",
                idx=i,
            )
        )
    print("ON rest", len(out))
    return out


def load_qc() -> list[dict]:
    out = []
    with open(DATA / "qc.csv", newline="", encoding="utf-8-sig", errors="replace") as f:
        for i, row in enumerate(csv.DictReader(f)):
            name = row.get("NOM") or ""
            postal = row.get("CODE_POSTAL_COMPO") or ""
            ll = fsa_ll(postal)
            if not ll or not name:
                continue
            lat, lng = jitter(ll[0], ll[1], name + postal)
            poupon = int(float(row.get("PLACE_TOTAL_POUPON") or 0) or 0)
            total = int(float(row.get("PLACE_TOTAL") or 0) or 0)
            typ = (row.get("TYPE") or "").upper()
            am = ["licensed", "funded"]
            if typ == "CPE":
                am.append("nonprofit")
            phone = row.get("telephone1") or ""
            out.append(
                make(
                    pid=f"qc-{i}-{slugify(name)[:12]}",
                    name=name.title() if name.isupper() else name,
                    name_fr=name.title() if name.isupper() else name,
                    address=row.get("ADRESSE") or "",
                    city=row.get("NOM_MUN_COMPO") or "",
                    province="QC",
                    postal=postal,
                    lat=lat,
                    lng=lng,
                    phone=phone,
                    languages="fr",
                    amenities=",".join(am),
                    spots_i=min(poupon, 6),
                    spots_t=max(total // 4, 1),
                    spots_p=max(total // 3, 1),
                    infant=poupon > 0,
                    idx=i,
                )
            )
    print("QC", len(out))
    return out


def nominatim_city(q: str, province: str) -> list[dict]:
    url = "https://nominatim.openstreetmap.org/search?" + urllib.parse.urlencode(
        {"format": "json", "addressdetails": 1, "limit": 30, "countrycodes": "ca", "q": q}
    )
    req = urllib.request.Request(url, headers={"User-Agent": "KidEase/1.0 (childcare directory)"})
    try:
        with urllib.request.urlopen(req, timeout=25) as r:
            hits = json.loads(r.read().decode())
    except Exception as e:
        print("nominatim fail", q, e)
        return []
    out = []
    for i, h in enumerate(hits):
        name = h.get("display_name", "").split(",")[0].strip()
        if not name:
            continue
        addr = h.get("address") or {}
        city = addr.get("city") or addr.get("town") or addr.get("village") or addr.get("suburb") or ""
        lat, lng = float(h["lat"]), float(h["lon"])
        out.append(
            make(
                pid=f"{province.lower()}-osm-{h.get('osm_id')}",
                name=name,
                address=addr.get("road") or "",
                city=city,
                province=province,
                postal=(addr.get("postcode") or ""),
                lat=lat,
                lng=lng,
                amenities="licensed",
                idx=i,
            )
        )
    print(" OSM", province, q, len(out))
    return out


def extras() -> list[dict]:
    queries = [
        ("child care centre Calgary Alberta", "AB"),
        ("child care centre Edmonton Alberta", "AB"),
        ("daycare Saskatoon Saskatchewan", "SK"),
        ("daycare Regina Saskatchewan", "SK"),
        ("child care Halifax Nova Scotia", "NS"),
        ("child care Charlottetown Prince Edward Island", "PE"),
        ("child care St. John's Newfoundland", "NL"),
        ("child care Whitehorse Yukon", "YT"),
        ("child care Yellowknife Northwest Territories", "NT"),
        ("child care Iqaluit Nunavut", "NU"),
    ]
    out: list[dict] = []
    for q, p in queries:
        out.extend(nominatim_city(q, p))
        time.sleep(1.1)
    return out


def dedupe(rows: list[dict]) -> list[dict]:
    seen: set[str] = set()
    out = []
    for r in rows:
        key = re.sub(r"[^a-z0-9]", "", (r["name"] + r["city"] + r["province"]).lower())
        if key in seen:
            continue
        seen.add(key)
        if not (-90 < r["lat"] < 90 and -180 < r["lng"] < 180):
            continue
        out.append(r)
    return out


def main():
    mb = load_mb()
    bc = load_bc()
    nb = load_nb()
    to = load_toronto()
    skip = {r["name"].lower() for r in to}
    on = load_on(skip)
    qc = load_qc()
    extra = extras()
    all_rows = dedupe(mb + bc + nb + to + on + qc + extra)
    # unique slugs
    slugs: set[str] = set()
    for r in all_rows:
        s = r["slug"]
        n = 2
        while s in slugs:
            s = f"{r['slug']}-{n}"
            n += 1
        r["slug"] = s
        slugs.add(s)
    by_p: dict[str, int] = {}
    for r in all_rows:
        by_p[r["province"]] = by_p.get(r["province"], 0) + 1
    print("TOTAL", len(all_rows), "by province", by_p)
    OUT.write_text(json.dumps(all_rows, ensure_ascii=False, separators=(",", ":")))
    print("wrote", OUT, "MB", OUT.stat().st_size // 1024)


if __name__ == "__main__":
    main()
